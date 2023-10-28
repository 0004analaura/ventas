import os
import openpyxl
import sys
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders




# ============ Productos ============

def crearProductos(codigo, nombre, existencia, proveedor, precio):
    def comprobarProducto(codigo):
        if os.path.exists("productos.txt"):
            with open("productos.txt", "r") as productos:
                for producto in productos:
                    if producto.split("|")[0] == codigo:
                        return True
        return False
    if comprobarProducto(codigo):
        print("El producto ", codigo, " ya existe")
    else:
        with open("productos.txt", "a") as productos:
            productos.write(codigo + "|" + nombre + "|" +
                            existencia + "|" + proveedor + "|" + precio + "\n")
        print("Producto " , codigo, " agregado")


def listarProductos():
    # Leer el archivo productos.txt
    productos = open("productos.txt", "r")
    print("====== Productos ======")
    for producto in productos:
        print(producto)
    print("=======================")
    productos.close()


def actualizarProductos(codigo, nombre, existencia, proveedor, precio):
    # Actualizar los datos parametrados en un archivo productos.txt
    productos = open("productos.txt", "r")
    productosTemp = open("productosTemp.txt", "w")
    for producto in productos:
        if producto.split("|")[0] == codigo:
            productosTemp.write(codigo + "|" + nombre + "|" +
                                existencia + "|" + proveedor + "|" + precio + "\n")
        else:
            productosTemp.write(producto)
            print("Producto ", codigo, " actualizado exitosamente")
    productos.close()
    productosTemp.close()
    os.remove("productos.txt")
    os.rename("productosTemp.txt", "productos.txt")


def actualizarExistencias(codigo, nuevaExistencia):
    # Actualizar las existencias de un producto
    productos = open("productos.txt", "r")
    productosTemp = open("productosTemp.txt", "w")
    for producto in productos:
        if producto.split("|")[0] == codigo:
            productosTemp.write(producto.split("|")[0] + "|" + producto.split("|")[
                                1] + "|" + nuevaExistencia + "|" + producto.split("|")[3] + "|" + producto.split("|")[4])
        else:
            productosTemp.write(producto)
            print("Existencia de producto ", codigo, " actualizada exitosamente a: ", nuevaExistencia )
    productos.close()
    productosTemp.close()
    os.remove("productos.txt")
    os.rename("productosTemp.txt", "productos.txt")


def eliminarProductos(codigo):
    # Eliminar los datos parametrados en un archivo productos.txt
    productos = open("productos.txt", "r")
    productosTemp = open("productosTemp.txt", "w")
    for producto in productos:
        if producto.split("|")[0] != codigo:
            productosTemp.write(producto)
            print("Producto ",codigo, " eliminado exitosamente")
    productos.close()
    productosTemp.close()
    os.remove("productos.txt")
    os.rename("productosTemp.txt", "productos.txt")

# ============ Clientes ============


def crearClientes(codigo, nombre, direccion, telefono):
    def comprobarCliente(codigo):
        # Crear el archivo si no existe
        with open("clientes.txt", "a") as f:
            pass

        with open("clientes.txt", "r") as clientes:
            for cliente in clientes:
                if cliente.split("|")[0] == codigo:
                    return True
        return False

    if comprobarCliente(codigo):
        print("El cliente" ,codigo, " ya existe")
    else:
        with open("clientes.txt", "a") as clientes:
            clientes.write(codigo + "|" + nombre + "|" +
                           direccion + "|" + telefono + "\n")
        print("Cliente " , codigo, " agregado")


def listarClientes():
    # Leer el archivo clientes.txt
    clientes = open("clientes.txt", "r")
    print("====== Clientes ======")
    for cliente in clientes:
        print(cliente)
    print("======================")
    clientes.close()


def editarCliente(codigo, nombre, direccion, telefono):
    # Actualizar los datos parametrados en un archivo clientes.txt
    clientes = open("clientes.txt", "r")
    clientesTemp = open("clientesTemp.txt", "w")
    for cliente in clientes:
        if cliente.split("|")[0] == codigo:
            clientesTemp.write(codigo + "|" + nombre + "|" +
                               direccion + "|" + telefono + "\n")
            print("Cliente ", codigo, " actualizado exitosamente")
        else:
            clientesTemp.write(cliente)
    clientes.close()
    clientesTemp.close()
    os.remove("clientes.txt")
    os.rename("clientesTemp.txt", "clientes.txt")


def eliminarCliente(codigo):
    # Eliminar los datos parametrados en un archivo clientes.txt
    clientes = open("clientes.txt", "r")
    clientesTemp = open("clientesTemp.txt", "w")
    for cliente in clientes:
        if cliente.split("|")[0] != codigo:
            clientesTemp.write(cliente)
            print("Cliente " ,codigo, " Eliminado exitosamente")
    clientes.close()
    clientesTemp.close()
    os.remove("clientes.txt")
    os.rename("clientesTemp.txt", "clientes.txt")

# ============ Ventas ============


def crearVentas(codigo, fecha, cliente, codigo_producto, cantidad):
    # Comprobar si hay existencias del producto y si hay crear la venta y descontar las existencias
    seHizoVenta = False  # Inicializamos seHizoVenta
    with open("productos.txt", "r") as productos:
        for producto in productos:
            if producto.split("|")[0] == codigo_producto:
                if int(producto.split("|")[2]) >= int(cantidad):
                    # Calcular total
                    total = float(producto.split("|")[4]) * int(cantidad)
                    with open("ventas.txt", "a") as ventas:
                        ventas.write(f"{codigo}|{fecha}|{cliente}|{codigo_producto}|{cantidad}|{total}\n")
                    seHizoVenta = True
                    print("Venta realizada con éxtio. Código de venta: " , codigo)
                    print("Producto: " , codigo_producto, " Cantidad: " , cantidad, " Total: " , total)
                    
                else:
                    print("No hay existencias suficientes del producto: " , codigo_producto)

    if seHizoVenta:
        actualizarExistencias(codigo_producto, str(int(producto.split("|")[2]) - int(cantidad)))


def listarVentas():
    # Leer el archivo ventas.txt
    ventas = open("ventas.txt", "r")
    print("====== Ventas ======")
    for venta in ventas:
        print(venta)
    print("====================")
    ventas.close()


def eliminarVenta(codigo):
    # Eliminar los datos parametrados en un archivo ventas.txt
    with open("ventas.txt", "r") as ventas:
        lines = ventas.readlines()
        
        venta_eliminada = None  # Guardar la venta eliminada
        
    # Buscar la venta a eliminar y revertir la cantidad de producto
    with open("ventas.txt", "w") as ventas:
        for line in lines:
            if line.split("|")[0] == codigo:
                venta_eliminada = line
            else:
                ventas.write(line)

    if venta_eliminada:
        # Revertir la cantidad de producto en el archivo de productos
        _, _, _, codigo_producto, cantidad, _ = venta_eliminada.split("|")
        cantidad = int(cantidad)
        actualizarExistencias(codigo_producto, str(int(getCantidadProducto(codigo_producto)) + cantidad))
        print("Venta", codigo, "anulada exitosamente.")
    else:
        print("No se encontró la venta con el código", codigo)
        
        # Función para obtener la cantidad de producto a partir del código del producto
def getCantidadProducto(codigo_producto):
    with open("productos.txt", "r") as productos:
        for producto in productos:
            if producto.split("|")[0] == codigo_producto:
                return producto.split("|")[2].strip()
    return None


# Exportar ventas a un archivo de excel (.xlsx)


def reporteVentasCliente(codigo_cliente, correo):
    # Obtener ventas por codigo de cliente y guardarlas en un excel y enviar el excel por correo
    workbook = openpyxl.Workbook()
   # workbook.save("ventasCliente.xlsx")
    worksheet= workbook.worksheets[0]
    # worksheet = workbook.create_sheet("Reporte")
   # estilo = workbook.add_format({'bold': True, 'font_color': 'white', 'bg_color': 'blue'})
    worksheet["A1"]="Codigo"
    worksheet["B1"]="Fecha"
    worksheet["C1"]="Codigo producto"
    worksheet["D1"]="Codigo cliente"
    worksheet["E1"]="Cantidad productos"
    worksheet["F1"]="Total venta"


    # worksheet.write('B1', 'Fecha', estilo)
    # worksheet.write('C1', 'Código producto', estilo)
    # worksheet.write('D1', 'Código cliente', estilo)
    # worksheet.write('E1', 'Cantidad productos', estilo)
    # worksheet.write('F1', 'Total venta', estilo)
    columnas=("A", "B", "C", "D", "E", "F")
    row = 2
    with open("ventas.txt", "r") as archivoVenta:
        ventas = archivoVenta.readlines()
    for venta in ventas:

        datos=venta.split("|")
        if(datos[2]!=codigo_cliente):
            continue
        worksheet.insert_rows(row)
        for i  in range (len(datos)):

            col=columnas[i] +str(row)
            print(col + str(row))
            worksheet[col]=datos[i]

        #if venta.split("|")[2] == codigo_cliente:
         #   worksheet.write(row, 0, venta.split("|")[0])
          #  worksheet.write(row, 1, venta.split("|")[1])
           # worksheet.write(row, 2, venta.split("|")[2])
            #worksheet.write(row, 3, venta.split("|")[3])
           # worksheet.write(row, 4, venta.split("|")[4])
            #worksheet.write(row, 5, venta.split("|")[5])
        row += 1

    workbook.save("ventasCliente.xlsx")
    workbook.close()
    
    # Enviar el archivo por correo

    #conexion a smtp
    conexion = smtplib.SMTP('smtp.gmail.com', 587)
    conexion.ehlo()
    
    #conexion segura TLS
    conexion.starttls() 
    
    # Crear variables necesarias para enviar
    USUARIO = os.getenv("Email_user")
    CONTRASENA = os.getenv("Email_secret")
    
    #login
    conexion.login(USUARIO, CONTRASENA)

    # Configuración del correo
    asunto = "Reporte de ventas por cliente"    
    mensaje = MIMEMultipart()
    mensaje['From'] = USUARIO
    mensaje['To'] = correo
    mensaje['Subject'] = asunto
     
    cuerpo = "El reporte de ventas por cliente se encuentra adjunto en este correo"
    
    # Agrega el cuerpo
    mensaje.attach(MIMEText(cuerpo, 'plain'))
    
    adjunto_path = "ventasCliente.xlsx"
    
    if adjunto_path:
        # Agrega el adjunto
        with open(adjunto_path, 'rb') as adjunto_file:
            adjunto_parte = MIMEBase('application', 'octet-stream')
            adjunto_parte.set_payload(adjunto_file.read())
            encoders.encode_base64(adjunto_parte)
            adjunto_parte.add_header('Content-Disposition', f'attachment; filename="ventasCliente.xlsx"')
            mensaje.attach(adjunto_parte)
    
    # Enviar el correo
    conexion.sendmail(USUARIO, correo, mensaje.as_string())

    print("Correo enviado a:", correo, "con el reporte de ventas por cliente")
    conexion.quit()



def reporteVentasProducto(codigo_producto, correo):
     # Obtener ventas por codigo de cliente y guardarlas en un excel y enviar el excel por correo
    workbook = openpyxl.Workbook()
   # workbook.save("ventasProducto.xlsx")
    worksheet= workbook.worksheets[0]
    # worksheet = workbook.create_sheet("Reporte")
   # estilo = workbook.add_format({'bold': True, 'font_color': 'white', 'bg_color': 'blue'})
    worksheet["A1"]="Codigo"
    worksheet["B1"]="Fecha"
    worksheet["C1"]="Codigo producto"
    worksheet["D1"]="Codigo cliente"
    worksheet["E1"]="Cantidad productos"
    worksheet["F1"]="Total venta"


    # worksheet.write('B1', 'Fecha', estilo)
    # worksheet.write('C1', 'Código producto', estilo)
    # worksheet.write('D1', 'Código cliente', estilo)
    # worksheet.write('E1', 'Cantidad productos', estilo)
    # worksheet.write('F1', 'Total venta', estilo)
    columnas=("A", "B", "C", "D", "E", "F")
    row = 3
    with open("ventas.txt", "r") as archivoVenta:
        ventas = archivoVenta.readlines()
    for venta in ventas:

        datos=venta.split("|")
        if(datos[3]!=codigo_producto):
            continue
        worksheet.insert_rows(row)
        for i  in range (len(datos)):

            col=columnas[i] +str(row)
            print(col + str(row))
            worksheet[col]=datos[i]

        #if venta.split("|")[2] == codigo_cliente:
         #   worksheet.write(row, 0, venta.split("|")[0])
          #  worksheet.write(row, 1, venta.split("|")[1])
           # worksheet.write(row, 2, venta.split("|")[2])
            #worksheet.write(row, 3, venta.split("|")[3])
           # worksheet.write(row, 4, venta.split("|")[4])
            #worksheet.write(row, 5, venta.split("|")[5])
        row += 1

    workbook.save("ventasProducto.xlsx")
    workbook.close()
    
    
    # Enviar el archivo por correo

    #conexion a smtp
    conexion = smtplib.SMTP('smtp.gmail.com', 587)
    conexion.ehlo()
    
    #conexion segura TLS
    conexion.starttls() 
    
    # Crear variables necesarias para enviar
    USUARIO = os.getenv("Email_user")
    CONTRASENA = os.getenv("Email_secret")
    
    #login
    conexion.login(USUARIO, CONTRASENA)

    # Configuración del correo
    asunto = "Reporte de ventas por producto"    
    mensaje = MIMEMultipart()
    mensaje['From'] = USUARIO
    mensaje['To'] = correo
    mensaje['Subject'] = asunto
     
    cuerpo = "El reporte de ventas por producto se encuentra adjunto en este correo"
    
    # Agrega el cuerpo
    mensaje.attach(MIMEText(cuerpo, 'plain'))
    
    adjunto_path = "ventasProducto.xlsx"
    
    if adjunto_path:
        # Agrega el adjunto
        with open(adjunto_path, 'rb') as adjunto_file:
            adjunto_parte = MIMEBase('application', 'octet-stream')
            adjunto_parte.set_payload(adjunto_file.read())
            encoders.encode_base64(adjunto_parte)
            adjunto_parte.add_header('Content-Disposition', f'attachment; filename="ventasProducto.xlsx"')
            mensaje.attach(adjunto_parte)
    
    # Enviar el correo
    conexion.sendmail(USUARIO, correo, mensaje.as_string())

    print("Correo enviado a:", correo, "con el reporte de ventas por producto")
    conexion.quit()

# ==================== Menu ====================
# Crear un menu con todas las funciones y un submenu con las funciones de cada seccion
def menu():
    print("""

==================== Menu ====================""")
    print("1. Clientes")
    print("2. Productos")
    print("3. Ventas")
    print("4. Reportes")
    print("5. Salir")
    print("""
=============================================""")
    opcion = input("Ingrese una opcion: ")
    if opcion == "1":
        menuClientes()
    elif opcion == "2":
        menuProductos()
    elif opcion == "3":
        menuVentas()
    elif opcion == "4":
        menuReportes()
    elif opcion == "5":
        print("""

Gracias por usar el sistema

""")
        exit()
    else:
        print("""

Opcion no valida

""")
        menu()

# ==================== Menu Clientes ====================
# Crear un submenu con las funciones de clientes
def menuClientes():
    print("""

==================== Clientes ====================""")
    print("1. Agregar cliente")
    print("2. Modificar cliente")
    print("3. Eliminar cliente")
    print("4. Listar clientes")
    print("5. Regresar")
    print("""

=============================================""")
    opcion = input("Ingrese una opcion: ")
    if opcion == "1":
        # codigo, nombre, direccion, telefono
        codigo = input("Ingrese el codigo del cliente: ")
        nombre = input("Ingrese el nombre del cliente: ")
        direccion = input("Ingrese la direccion del cliente: ")
        telefono = input("Ingrese el telefono del cliente: ")
        crearClientes(codigo, nombre, direccion, telefono)
        menuClientes()
    elif opcion == "2":
        # codigo, nombre, direccion, telefono
        codigo = input("Ingrese el codigo del cliente a editar: ")
        nombre = input("Ingrese el nuevo nombre del cliente a editar: ")
        direccion = input("Ingrese la nueva direccion del cliente a editar: ")
        telefono = input("Ingrese el nuevo telefono del cliente a editar: ")
        editarCliente(codigo, nombre, direccion, telefono)
        menuClientes()
    elif opcion == "3":
        codigo = input("Ingrese el codigo del cliente a eliminar: ")
        eliminarCliente(codigo)
        menuClientes()
    elif opcion == "4":
        listarClientes()
        menuClientes()
    elif opcion == "5":
        menu()
    else:
        print("""

Opcion no valida

""")
        menuClientes()

# ==================== Menu Productos ====================
# Crear un submenu con las funciones de productos
def menuProductos():
    print("""

==================== Productos ====================""")
    print("1. Agregar producto")
    print("2. Actualizar existencias")
    print("3. Modificar producto")
    print("4. Eliminar producto")
    print("5. Listar productos")
    print("6. Regresar")
    print("""

=============================================""")
    opcion = input("Ingrese una opcion: ")
    if opcion == "1":
        # codigo, nombre, existencia, proveedor, precio
        codigo = input("Ingrese el codigo del producto: ")
        nombre = input("Ingrese el nombre del producto: ")
        existencia = input("Ingrese la existencia del producto: ")
        proveedor = input("Ingrese el proveedor del producto: ")
        precio = input("Ingrese el precio del producto: ")
        crearProductos(codigo, nombre, existencia, proveedor, precio)
        menuProductos()
    elif opcion == "2":
        # codigo, nuevaExistencia
        codigo = input("Ingrese el codigo del producto a actualizar: ")
        nuevaExistencia = input("Ingrese la nueva existencia del producto a actualizar: ")
        actualizarExistencias(codigo, nuevaExistencia)
        menuProductos()
    elif opcion == "3":
        # codigo, nombre, existencia, proveedor, precio
        codigo = input("Ingrese el codigo del producto a actualizar: ")
        nombre = input("Ingrese el nuevo nombre del producto a actualizar: ")
        existencia = input("Ingrese la nueva existencia del producto a actualizar: ")
        proveedor = input("Ingrese el nuevo proveedor del producto a actualizar: ")
        precio = input("Ingrese el nuevo precio del producto a actualizar: ")
        actualizarProductos(codigo, nombre, existencia, proveedor, precio)
        menuProductos()
    elif opcion == "4":
        codigo = input("Ingrese el codigo del producto a eliminar: ")
        eliminarProductos(codigo)
        menuProductos()
    elif opcion == "5":
        listarProductos()
        menuProductos()
    elif opcion == "6":
        menu()
    else:
        print("""

Opcion no valida

""")
        menuProductos()

# ==================== Menu Ventas ====================
# Crear un submenu con las funciones de ventas
def menuVentas():
    print("""

==================== Ventas ====================""")
    print("1. Agregar venta")
    print("2. Anular venta")
    print("3. Listar ventas")
    print("4. Regresar")
    print("""

=============================================""")
    opcion = input("Ingrese una opcion: ")
    if opcion == "1":
        codigo = input("Ingrese el codigo de la venta: ")
        fecha = input("Ingrese la fecha de la venta: ")
        cliente = input("Ingrese el codigo del cliente de la venta: ")
        producto = input("Ingrese el codigo del producto de la venta: ")
        cantidad = input("Ingrese la cantidad del producto a la venta: ")
        crearVentas(codigo, fecha, cliente, producto, cantidad)
        menuVentas()
    elif opcion == "2":
        codigo = input("Ingrese el codigo de la venta a eliminar: ")
        eliminarVenta(codigo)
        menuVentas()
    elif opcion == "3":
        listarVentas()
        menuVentas()
    elif opcion == "4":
        menu()
    else:
        print("""

Opcion no valida

""")
        menuVentas()

# ==================== Menu Reportes ====================
# Crear un submenu con las funciones de reportes
def menuReportes():
    print("""

==================== Reportes ====================""")
    print("1. Reporte de ventas por cliente")
    print("2. Reporte de ventas por producto")
    print("3. Regresar")
    print("""

=============================================""")
    opcion = input("Ingrese una opcion: ")
    if opcion == "1":
        codigo = input("Ingrese el codigo del cliente: ")
        correo = input("Ingrese el correo receptor: ")
        reporteVentasCliente(codigo, correo)
        menuReportes()
    elif opcion == "2":
        codigo = input("Ingrese el codigo del producto: ")
        correo = input("Ingrese el correo receptor: ")
        reporteVentasProducto(codigo, correo)
        menuReportes()
    elif opcion == "3":
        menu()
    else:
        print("""

Opcion no valida

""")
        menuReportes()

# ==================== Menu Argumentos ====================
# Crear un menu con todas las funciones parametrando con argumentos
def menuArgumentos():
    # Menu Argumentos:
    # 1. Productos
    # 2. Clientes
    # 3. Ventas
    # 4. Reportes
    # 5. Ayuda
    # 1.1. Crear
    # 1.2. Actualizar Existencias
    # 1.3. Actualizar
    # 1.4. Listar
    # 1.5. Eliminar
    # 2.1. Crear
    # 2.2. Actualizar
    # 2.3. Listar
    # 2.4. Eliminar
    # 3.1. Crear
    # 3.2. Anular
    # 3.3. Listar
    # 4.1. Reporte de ventas por cliente
    # 4.2. Reporte de ventas por producto
    opcionElegida = sys.argv[1]
    if opcionElegida.lower() == "productos":
        opcionElegida = sys.argv[2]
        if opcionElegida.lower() == "crear":
            codigo = sys.argv[3]
            nombre = sys.argv[4]
            existencia = sys.argv[5]
            proveedor = sys.argv[6]
            precio = sys.argv[7]
            crearProductos(codigo, nombre, existencia, proveedor, precio)
        elif opcionElegida.lower() == "actualizarExistencias":
            codigo = sys.argv[3]
            nuevaExistencia = sys.argv[4]
            actualizarExistencias(codigo, nuevaExistencia)
        elif opcionElegida.lower() == "actualizar":
            codigo = sys.argv[3]
            nombre = sys.argv[4]
            existencia = sys.argv[5]
            proveedor = sys.argv[6]
            precio = sys.argv[7]
            actualizarProductos(codigo, nombre, existencia, proveedor, precio)
        elif opcionElegida.lower() == "listar":
            listarProductos()
        elif opcionElegida.lower() == "eliminar":
            codigo = sys.argv[3]
            eliminarProductos(codigo)
        else:
            print("""

Opcion no valida

""")
    elif opcionElegida.lower() == "clientes":
        opcionElegida = sys.argv[2]
        if opcionElegida.lower() == "crear":
            # codigo, nombre, direccion, telefono
            codigo = sys.argv[3]
            nombre = sys.argv[4]
            direccion = sys.argv[5]
            telefono = sys.argv[6]
            crearClientes(codigo, nombre, direccion, telefono)
        elif opcionElegida.lower() == "actualizar":
            # codigo, nombre, direccion, telefono
            codigo = sys.argv[3]
            nombre = sys.argv[4]
            direccion = sys.argv[5]
            telefono = sys.argv[6]
            editarCliente(codigo, nombre, direccion, telefono)
        elif opcionElegida.lower() == "listar":
            listarClientes()
        elif opcionElegida.lower() == "eliminar":
            codigo = sys.argv[3]
            eliminarCliente(codigo)
        else:
            print("""

Opcion no valida

""")
    elif opcionElegida.lower() == "ventas":
        opcionElegida = sys.argv[2]
        if opcionElegida.lower() == "crear":
            # codigo, fecha, cliente, producto, cantidad
            codigo = sys.argv[3]
            fecha = sys.argv[4]
            cliente = sys.argv[5]
            producto = sys.argv[6]
            cantidad = sys.argv[7]
            crearVentas(codigo, fecha, cliente, producto, cantidad)
        elif opcionElegida.lower() == "anular":
            codigo = sys.argv[3]
            eliminarVenta(codigo)
        elif opcionElegida.lower() == "listar":
            listarVentas()
        else:
            print("""

Opcion no valida

""")
    elif opcionElegida.lower() == "reportes":
        opcionElegida = sys.argv[2]
        if opcionElegida.lower() == "cliente":
            codigo = sys.argv[3]
            correo = sys.argv[4]
            reporteVentasCliente(codigo, correo)
        elif opcionElegida.lower() == "producto":
            codigo = sys.argv[3]
            correo = sys.argv[4]
            reporteVentasProducto(codigo, correo)
        else:
            print("""

Opcion no valida

""")
    elif opcionElegida.lower() == "ayuda":
        print("""
        Comandos:
        python main.py productos crear "Codigo del producto" "Nombre del producto" "Existencias" "Proveedor" "Precio"
        python main.py productos actualizarExistencias "Codigo del producto" "Nueva existencia"
        python main.py productos actualizar "Codigo del producto" "Nombre del producto" "Existencias" "Proveedor" "Precio"
        python main.py productos listar
        python main.py productos eliminar "Codigo del producto"
        python main.py clientes crear "Codigo del cliente" "Nombre del cliente" "Direccion del cliente" "Telefono del cliente"
        python main.py clientes actualizar "Codigo del cliente" "Nombre del cliente" "Direccion del cliente" "Telefono del cliente"
        python main.py clientes listar
        python main.py clientes eliminar "Codigo del cliente"
        python main.py ventas crear "Codigo de la venta" "Fecha de la venta" "Codigo del cliente" "Codigo del producto" "Cantidad"
        python main.py ventas anular "Codigo de la venta"
        python main.py ventas listar
        python main.py reportes cliente "Codigo del cliente" "Correo del cliente"
        python main.py reportes producto "Codigo del producto" "Correo del cliente"
        """)
        


def principal():
    if not len(sys.argv) >= 2:
        menu()
    else:
        menuArgumentos()

principal()